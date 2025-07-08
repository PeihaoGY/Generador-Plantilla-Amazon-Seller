#Generador de plantilla para Amazon Seller
#Esta aplicacion generara una plantilla siguiendo las pautas de Amazon Seller para cambiar precios e inventario.
#El proceso utiliza 2 fuentes de informacion: "stocks.csv" (fichero de productos diaros a subir) y la plantilla de Amazon.
#Version 1.0
#Fecha: 08/07/2025
#Autor: Peihao Guo Yang
#Creado para la empresa: DIRAC DIST S.L.

import pandas as pd
import traceback
from openpyxl import load_workbook

# ---------- CONFIGURACIÓN ----------
CSV_DIARIO = r'\\RED\Ruta\ejemplo\donde\esta\todo\stocks.csv'
TEMPLATE_PATH = r'\\RED\Ruta\ejemplo\donde\esta\todo\plantillaEjemplo.xlsx'
OUTPUT_DIRECTORY = r'\\RED\Ruta\ejemplo\donde\esta\todo\ArchivosFinales'
OUTPUT_XLSX = f'{OUTPUT_DIRECTORY}\\Resultado_Amazon.xlsx'
OUTPUT_TXT = f'{OUTPUT_DIRECTORY}\\Resultado_Amazon.txt'

try:
    df_csv = pd.read_csv(CSV_DIARIO, sep='\t')
    df_csv.columns = [col.strip() for col in df_csv.columns]

    # Crear columnas requeridas
    df_csv['product-id'] = df_csv['EAN']
    df_csv['product-id-type'] = 'EAN'

    # Filtrar columnas para plantilla
    df_resultado = df_csv[['SKU', 'product-id', 'product-id-type', 'Stock']]
    df_resultado.columns = ['sku', 'product-id', 'product-id-type', 'quantity']

    # ----- Guardar en Excel usando plantilla -----
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb['Plantilla']

    # Insertar datos desde la segunda fila (asumiendo encabezados ya en la plantilla)
    start_row = 2
    for i, row in df_resultado.iterrows():
        ws.cell(row=start_row + i, column=1).value = row['sku']
        ws.cell(row=start_row + i, column=2).value = row['product-id']
        ws.cell(row=start_row + i, column=3).value = row['product-id-type']
        ws.cell(row=start_row + i, column=4).value = row['quantity']

    # Guardar como nuevo Excel
    wb.save(OUTPUT_XLSX)
    print(f"✅ Archivo Excel generado: {OUTPUT_XLSX}")

    # ----- Guardar como TXT -----
    df_resultado.to_csv(OUTPUT_TXT, sep='\t', index=False)
    print(f"✅ Archivo TXT generado: {OUTPUT_TXT}")
except Exception as e:
    with open("error_log.txt", "w") as f:
        traceback.print_exc(file=f)